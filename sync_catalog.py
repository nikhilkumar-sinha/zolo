import os
import sys
import sqlite3
import json
import datetime
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_FILE = os.path.join(os.path.dirname(__file__), 'zolofresh.db')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'catalog.xlsx')
JS_FILE = os.path.join(os.path.dirname(__file__), 'catalog-data.js')

# Predefined categories mapping to initialize the database
DEFAULT_CATEGORIES = [
    {"id": 1, "category_name": "Fruits & Fresh Produce", "slug": "fruits-fresh-produce", "parent_id": None, "status": 1},
    {"id": 2, "category_name": "Staples & Grains", "slug": "staples-grains", "parent_id": None, "status": 1},
    {"id": 3, "category_name": "Superfoods & Snacks", "slug": "superfoods-snacks", "parent_id": None, "status": 1},
    {"id": 4, "category_name": "Handlooms & Handicrafts", "slug": "handlooms-handicrafts", "parent_id": None, "status": 1},
    {"id": 5, "category_name": "Puja Essentials & Kits", "slug": "puja-essentials-kits", "parent_id": None, "status": 1},
    {"id": 6, "category_name": "Shahi Litchi", "slug": "shahi-litchi", "parent_id": 1, "status": 1},
    {"id": 7, "category_name": "Jardalu Mango", "slug": "jardalu-mango", "parent_id": 1, "status": 1},
    {"id": 8, "category_name": "Katarni Rice", "slug": "katarni-rice", "parent_id": 2, "status": 1},
    {"id": 9, "category_name": "Bhagalpuri Rice", "slug": "bhagalpuri-rice", "parent_id": 2, "status": 1},
    {"id": 10, "category_name": "Chana Ka Sattu", "slug": "chana-ka-sattu", "parent_id": 2, "status": 1},
    {"id": 11, "category_name": "Chura / Poha", "slug": "chura-poha", "parent_id": 2, "status": 1},
    {"id": 12, "category_name": "Mithila Makhana", "slug": "mithila-makhana", "parent_id": 3, "status": 1},
    {"id": 13, "category_name": "Bhuna Makhana", "slug": "bhuna-makhana", "parent_id": 3, "status": 1},
    {"id": 14, "category_name": "Silao Khaja", "slug": "silao-khaja", "parent_id": 3, "status": 1},
    {"id": 15, "category_name": "Tilkut", "slug": "tilkut", "parent_id": 3, "status": 1},
    {"id": 16, "category_name": "Thekua", "slug": "thekua", "parent_id": 3, "status": 1},
    {"id": 17, "category_name": "Jaggery", "slug": "jaggery", "parent_id": 3, "status": 1},
    {"id": 18, "category_name": "Traditional Sweets", "slug": "traditional-sweets", "parent_id": 3, "status": 1},
    {"id": 19, "category_name": "Madhubani Paintings", "slug": "madhubani-paintings", "parent_id": 4, "status": 1},
    {"id": 20, "category_name": "Bhagalpuri Silk", "slug": "bhagalpuri-silk", "parent_id": 4, "status": 1},
    {"id": 21, "category_name": "Sikki Grass Craft", "slug": "sikki-grass-craft", "parent_id": 4, "status": 1},
    {"id": 22, "category_name": "Diwali Puja Kit", "slug": "diwali-puja-kit", "parent_id": 5, "status": 1},
    {"id": 23, "category_name": "Durga Puja Kit", "slug": "durga-puja-kit", "parent_id": 5, "status": 1},
    {"id": 24, "category_name": "Satyanarayan Puja Kit", "slug": "satyanarayan-puja-kit", "parent_id": 5, "status": 1}
]

# Mapping between spreadsheet Category names and SQLite subcategory IDs / parent categories
PRODUCT_SUBCAT_MAPPING = {
    "shahi-litchi": 6,
    "jardalu-mango": 7,
    "organic-jamun": 1,
    "alphonso-mango": 7,
    "katarni-rice": 8,
    "bhagalpuri-rice": 9,
    "chana-sattu": 10,
    "chura-poha": 11,
    "darjeeling-tea": 8,
    "mithila-makhana": 12,
    "bhuna-makhana": 13,
    "makhana-kheer-kit": 18,
    "kashmiri-kesar": 12,
    "diwali-puja-package": 22,
    "durgapuja-package": 23,
    "satyanarayan-puja-package": 24,
    "silao-khaja": 14,
    "gaya-sesame-tilkut": 15,
    "chhath-thekua": 16,
    "sesame-anarsa": 18,
    "organic-jaggery": 17,
    "traditional-sweets": 18,
    "madhubani-paintings": 19,
    "bhagalpuri-silk": 20,
    "sikki-grass-craft": 21,
    "mysore-sandalwood": 20,
    "mithila-painting": 19
}

# Source of truth products compiled to match the screenshot exactly
INITIAL_PRODUCTS = [
    {"category": "Crafts", "name": "Madhubani Paintings (GI Tagged)", "sub_cat": "GI Tagged", "sku": "ZF-CR-020", "discount": 5, "price": 617.5, "sales_price": 650, "unit": "Handmade Canvas (A4 Size)", "stock": "50 Units", "image": ".\\assets\\mithila_makhana.jpg"},
    {"category": "Crafts", "name": "Bhagalpuri Silk (Tussar Stole)", "sub_cat": "GI Tagged", "sku": "ZF-CR-021", "discount": 5, "price": 712.5, "sales_price": 750, "unit": "Premium Silk Stole (2 Meters)", "stock": "50 Units", "image": ".\\assets\\katarni_rice.jpg"},
    {"category": "Crafts", "name": "Sikki Grass Craft Basket", "sub_cat": "GI Tagged", "sku": "ZF-CR-022", "discount": 5, "price": 275.5, "sales_price": 290, "unit": "Handwoven Utility Box", "stock": "50 Units", "image": ".\\assets\\puja_box.jpg"},
    {"category": "Crafts", "name": "Mithila Painting", "sub_cat": "GI Tagged", "sku": "ZF-CR-023", "discount": 5, "price": 569.05, "sales_price": 599, "unit": "Utility Box", "stock": "50 Units", "image": ".\\assets\\mithila_painting.jpg"},
    {"category": "Crafts", "name": "Pure Mysore Sandalwood Oil", "sub_cat": "GI Tagged", "sku": "ZF-CR-025", "discount": 5, "price": 902.5, "sales_price": 950, "unit": "Bottle of 5ml", "stock": "50 Units", "image": ".\\assets\\puja_box.jpg"},
    {"category": "Fruits", "name": "Certified Shahi Litchi", "sub_cat": "GI Tagged", "sku": "ZF-FR-001", "discount": 5, "price": 228.0, "sales_price": 240, "unit": "Per Kg (approx. 40-45 units)", "stock": "50 Kg", "image": ".\\assets\\shahi_litchi.jpg"},
    {"category": "Fruits", "name": "Bhagalpur Jardalu Mango", "sub_cat": "GI Tagged", "sku": "ZF-FR-002", "discount": 5, "price": 180.5, "sales_price": 190, "unit": "Per Kg (approx. 4-5 units)", "stock": "50 Kg", "image": ".\\assets\\jardalu_mango.jpg"},
    {"category": "Fruits", "name": "Wild Organic Jamun (Black Plum)", "sub_cat": "Organic", "sku": "ZF-FR-003", "discount": 0, "price": 180.0, "sales_price": 180, "unit": "Pack of 500g", "stock": "100 Grams", "image": ".\\assets\\Sweets.avif"},
    {"category": "Fruits", "name": "Devgad Alphonso Hapus Mangoes", "sub_cat": "GI Tagged", "sku": "ZF-FR-026", "discount": 5, "price": 807.5, "sales_price": 850, "unit": "Box of 6 Pieces", "stock": "50 Units", "image": ".\\assets\\jardalu_mango.jpg"},
    {"category": "Grains", "name": "Fragrant Katarni Rice", "sub_cat": "GI Tagged", "sku": "ZF-GR-007", "discount": 5, "price": 123.5, "sales_price": 130, "unit": "Pack of 1 Kg", "stock": "50 Kg", "image": ".\\assets\\katarni_rice.jpg"},
    {"category": "Grains", "name": "Bhagalpuri Rice", "sub_cat": "Organic", "sku": "ZF-GR-008", "discount": 5, "price": 104.5, "sales_price": 110, "unit": "Pack of 1 Kg", "stock": "50 Kg", "image": ".\\assets\\katarni_rice.jpg"},
    {"category": "Grains", "name": "Chana ka Sattu (Gram Sattu)", "sub_cat": "Organic", "sku": "ZF-GR-009", "discount": 5, "price": 85.5, "sales_price": 90, "unit": "Pack of 1 Kg", "stock": "50 Kg", "image": ".\\assets\\traditional_sattu.jpg"},
    {"category": "Grains", "name": "Chura / Poha / Avlaki", "sub_cat": "Organic", "sku": "ZF-GR-010", "discount": 5, "price": 76.0, "sales_price": 80, "unit": "Pack of 500g", "stock": "100 Grams", "image": ".\\assets\\katarni_rice.jpg"},
    {"category": "Grains", "name": "Darjeeling First Flush Black Tea", "sub_cat": "GI Tagged", "sku": "ZF-GR-024", "discount": 5, "price": 427.5, "sales_price": 450, "unit": "Pack of 250g", "stock": "100 Grams", "image": ".\\assets\\katarni_rice.jpg"},
    {"category": "Makhana", "name": "Premium Mithila Makhana", "sub_cat": "GI Tagged", "sku": "ZF-MA-004", "discount": 5, "price": 361.0, "sales_price": 380, "unit": "Pack of 500g (Jumbo Size)", "stock": "100 Grams", "image": ".\\assets\\mithila_makhana.jpg"},
    {"category": "Makhana", "name": "Bhuna Makhana (Roasted Foxnuts)", "sub_cat": "GI Tagged", "sku": "ZF-MA-005", "discount": 5, "price": 142.5, "sales_price": 150, "unit": "Pack of 200g", "stock": "100 Grams", "image": ".\\assets\\mithila_makhana.jpg"},
    {"category": "Makhana", "name": "Mithila Makhana Kheer Mix", "sub_cat": "GI Tagged", "sku": "ZF-MA-006", "discount": 5, "price": 237.5, "sales_price": 250, "unit": "Pack of 300g (Ready Pudding Mix)", "stock": "100 Grams", "image": ".\\assets\\mithila_makhana.jpg"},
    {"category": "Makhana", "name": "Grade A++ Kashmiri Kesar (Saffron)", "sub_cat": "GI Tagged", "sku": "ZF-MA-023", "discount": 5, "price": 332.5, "sales_price": 350, "unit": "Pack of 1 Gram", "stock": "100 Grams", "image": ".\\assets\\puja_box.jpg"},
    {"category": "Puja", "name": "Diwali Sacred Puja Package", "sub_cat": "Organic", "sku": "ZF-PU-017", "discount": 5, "price": 427.5, "sales_price": 450, "unit": "Complete Kit (12 items)", "stock": "50 Units", "image": ".\\assets\\puja_box.jpg"},
    {"category": "Puja", "name": "Durga Puja Sacred Package", "sub_cat": "Organic", "sku": "ZF-PU-018", "discount": 5, "price": 522.5, "sales_price": 550, "unit": "Complete Kit (15 items)", "stock": "50 Units", "image": ".\\assets\\puja_box.jpg"},
    {"category": "Puja", "name": "Satyanarayan Bhagwan Puja Package", "sub_cat": "Organic", "sku": "ZF-PU-019", "discount": 5, "price": 361.0, "sales_price": 380, "unit": "Complete Kit (10 items)", "stock": "50 Units", "image": ".\\assets\\puja_box.jpg"},
    {"category": "Sweets", "name": "Famous Silao Khaja (GI Tagged)", "sub_cat": "GI Tagged", "sku": "ZF-SW-011", "discount": 5, "price": 209.0, "sales_price": 220, "unit": "Pack of 500g (12 Pieces)", "stock": "100 Grams", "image": ".\\assets\\silao_khaja.jpg"},
    {"category": "Sweets", "name": "Gaya Special Sesame Tilkut", "sub_cat": "Organic", "sku": "ZF-SW-012", "discount": 5, "price": 266.0, "sales_price": 280, "unit": "Pack of 500g", "stock": "100 Grams", "image": ".\\assets\\gaya_tilkut.jpg"},
    {"category": "Sweets", "name": "Authentic Chhath Prasad Thekua", "sub_cat": "Organic", "sku": "ZF-SW-013", "discount": 5, "price": 228.0, "sales_price": 240, "unit": "Pack of 500g (Approx 16-18 pcs)", "stock": "100 Grams", "image": ".\\assets\\thekua.jpg"},
    {"category": "Sweets", "name": "Traditional Sesame Anarsa", "sub_cat": "Organic", "sku": "ZF-SW-014", "discount": 5, "price": 247.0, "sales_price": 260, "unit": "Pack of 500g", "stock": "100 Grams", "image": ".\\assets\\anarsa.jpg"},
    {"category": "Sweets", "name": "Organic Jaggery (Bheli)", "sub_cat": "Organic", "sku": "ZF-SW-015", "discount": 5, "price": 114.0, "sales_price": 120, "unit": "Pack of 1 Kg", "stock": "50 Gur", "image": ".\\assets\\puja_box.jpg"},
    {"category": "Sweets", "name": "Traditional Sweets", "sub_cat": "Traditional", "sku": "ZF-SW-016", "discount": 5, "price": 569.05, "sales_price": 599, "unit": "Pack of 1 Kg", "stock": "50 Kg", "image": ".\\assets\\Sweets.avif"}
]


def create_slug(name):
    # Standard slugify
    s = name.lower()
    s = re.sub(r'\(gi tagged\)', '', s)
    s = re.sub(r'\(saffron\)', '', s)
    s = re.sub(r'\(black plum\)', '', s)
    s = re.sub(r'\(gram sattu\)', '', s)
    s = re.sub(r'\(roasted foxnuts\)', '', s)
    s = re.sub(r'\(tussar stole\)', '', s)
    s = re.sub(r'\(bheli\)', '', s)
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s_]+', '-', s)
    s = s.strip('-')
    
    # Custom adjustments to match product mapping dictionary
    if 'litchi' in s: return 'shahi-litchi'
    if 'jardalu' in s: return 'jardalu-mango'
    if 'jamun' in s: return 'organic-jamun'
    if 'alphonso' in s: return 'alphonso-mango'
    if 'katarni' in s: return 'katarni-rice'
    if 'bhagalpuri-rice' in s: return 'bhagalpuri-rice'
    if 'sattu' in s: return 'chana-sattu'
    if 'poha' in s or 'chura' in s: return 'chura-poha'
    if 'kesar' in s: return 'kashmiri-kesar'
    if 'tea' in s: return 'darjeeling-tea'
    if 'kheer' in s: return 'makhana-kheer-kit'
    if 'bhuna-makhana' in s: return 'bhuna-makhana'
    if 'premium-mithila-makhana' in s or 'mithila-makhana' in s: return 'mithila-makhana'
    if 'diwali' in s: return 'diwali-puja-package'
    if 'durga' in s: return 'durgapuja-package'
    if 'satyanarayan' in s: return 'satyanarayan-puja-package'
    if 'khaja' in s: return 'silao-khaja'
    if 'tilkut' in s: return 'gaya-sesame-tilkut'
    if 'thekua' in s: return 'chhath-thekua'
    if 'anarsa' in s: return 'sesame-anarsa'
    if 'jaggery' in s: return 'organic-jaggery'
    if 'sandalwood' in s: return 'mysore-sandalwood'
    if 'silk' in s: return 'bhagalpuri-silk'
    if 'sikki' in s: return 'sikki-grass-craft'
    
    return s


def parse_stock_quantity(stock_str):
    if not stock_str:
        return 120
    # Extract digits
    digits = re.findall(r'\d+', str(stock_str))
    if digits:
        return int(digits[0])
    return 120


def clean_image_path(img_path):
    if not img_path:
        return ""
    # Convert .\assets\image.jpg to assets/image.jpg
    path = str(img_path).replace('\\', '/').strip()
    if path.startswith('./'):
        path = path[2:]
    elif path.startswith('.'):
        path = path[1:]
    if path.startswith('/'):
        path = path[1:]
    return path


def format_excel_sheet(ws):
    ws.views.sheetView[0].showGridLines = True
    
    # Theme color: HSL-aligned Premium Blue / Dark Forest Green
    header_fill = PatternFill(start_color="1A3D54", end_color="1A3D54", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Style headers
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_thin
    
    # Style cells
    data_font = Font(name="Segoe UI", size=10, color="333333")
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = border_thin
            
            # Formatting numbers
            if col in [1, 2, 3, 8, 10]:  # Text columns
                cell.alignment = align_left
            elif col in [4]:  # SKU
                cell.alignment = align_center
            elif col in [5, 6, 7, 9]:  # Numbers
                cell.alignment = align_center
                if col in [6, 7] and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

    # Auto adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)


def export_to_excel():
    print(f"Generating custom Excel catalog spreadsheet: {EXCEL_FILE}...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    
    headers = [
        "category", "Product Name", "Sub-Categorie", "SKU", 
        "Discount", "Price", "Sales Price", "Unit", "Stock Quantity", "Product Image"
    ]
    ws.append(headers)
    
    products_to_export = INITIAL_PRODUCTS
    
    # Try reading current SQLite state if DB exists and has records
    if os.path.exists(DB_FILE):
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT category_id, product_title, sku, price, unit, image, organic, gi_tagged 
                FROM products
            """)
            rows = cursor.fetchall()
            if rows:
                products_to_export = []
                for r in rows:
                    cat_id = r[0]
                    title = r[1]
                    sku = r[2]
                    sales_price = r[3]
                    unit = r[4]
                    image_path = r[5]
                    is_organic = r[6]
                    is_gi = r[7]
                    
                    # Resolve category name
                    cat_map = {1: "Fruits", 2: "Grains", 3: "Makhana", 4: "Crafts", 5: "Puja"}
                    # Sweets maps to category ID 14, 15, 16, 17, 18
                    if cat_id in [14, 15, 16, 17, 18]:
                        category = "Sweets"
                    elif cat_id in [12, 13]:
                        category = "Makhana"
                    elif cat_id in [8, 9, 10, 11]:
                        category = "Grains"
                    elif cat_id in [22, 23, 24]:
                        category = "Puja"
                    elif cat_id in [19, 20, 21]:
                        category = "Crafts"
                    else:
                        category = cat_map.get(cat_id, "Fruits")
                        
                    sub_cat = "GI Tagged" if is_gi else ("Organic" if is_organic else "Traditional")
                    discount = 5 if is_gi or is_organic else 0
                    price = round(sales_price * (1 - discount/100), 2)
                    
                    # Format image path back to .\assets\filename.ext
                    win_image = ".\\" + image_path.replace('/', '\\') if image_path else ""
                    
                    products_to_export.append({
                        "category": category,
                        "name": title,
                        "sub_cat": sub_cat,
                        "sku": sku,
                        "discount": discount,
                        "price": price,
                        "sales_price": sales_price,
                        "unit": unit,
                        "stock": "50 Units",
                        "image": win_image
                    })
            conn.close()
        except Exception as e:
            print(f"Note: Could not query products from DB during export, using bootstrap list. ({e})")
            
    for p in products_to_export:
        ws.append([
            p["category"], p["name"], p["sub_cat"], p["sku"],
            p["discount"], p["price"], p["sales_price"], p["unit"], p["stock"], p["image"]
        ])
        
    format_excel_sheet(ws)
    wb.save(EXCEL_FILE)
    print(f"Successfully generated Excel catalog sheet: {EXCEL_FILE}")


def import_from_excel():
    if not os.path.exists(EXCEL_FILE):
        print(f"Spreadsheet '{EXCEL_FILE}' not found. Running initial export to bootstrap catalog...")
        export_to_excel()
        return

    print(f"Reading Excel catalog spreadsheet: {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active  # Uses the active sheet (Products)
    
    products = []
    for row in range(2, ws.max_row + 1):
        category = ws.cell(row=row, column=1).value
        prod_name = ws.cell(row=row, column=2).value
        sub_cat = ws.cell(row=row, column=3).value
        sku = ws.cell(row=row, column=4).value
        discount = ws.cell(row=row, column=5).value
        price = ws.cell(row=row, column=6).value
        sales_price = ws.cell(row=row, column=7).value
        unit = ws.cell(row=row, column=8).value
        stock = ws.cell(row=row, column=9).value
        image = ws.cell(row=row, column=10).value
        
        if sku and prod_name and category:
            products.append({
                "category": str(category).strip(),
                "name": str(prod_name).strip(),
                "sub_cat": str(sub_cat).strip() if sub_cat else "Traditional",
                "sku": str(sku).strip(),
                "discount": float(discount) if discount is not None else 0.0,
                "price": float(price) if price is not None else 0.0,
                "sales_price": float(sales_price) if sales_price is not None else 0.0,
                "unit": str(unit).strip() if unit else "Unit",
                "stock": str(stock).strip() if stock else "120 Units",
                "image": str(image).strip() if image else ""
            })
            
    print(f"Loaded {len(products)} products from spreadsheet catalog.")
    
    # Update SQLite database
    print(f"Updating SQLite database: {DB_FILE}...")
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON;")
        
        # Clear existing tables safely
        cursor.execute("DELETE FROM products;")
        cursor.execute("DELETE FROM categories;")
        cursor.execute("DELETE FROM sqlite_sequence WHERE name IN ('products', 'categories');")
        
        # Seed the 24 categories
        for c in DEFAULT_CATEGORIES:
            cursor.execute("""
                INSERT INTO categories (id, category_name, slug, parent_id, status)
                VALUES (?, ?, ?, ?, ?)
            """, (c["id"], c["category_name"], c["slug"], c["parent_id"], c["status"]))
            
        # Parse products from spreadsheet and insert
        for p in products:
            slug = create_slug(p["name"])
            category_id = PRODUCT_SUBCAT_MAPPING.get(slug, 1)
            
            # Map flags based on Sub-Categorie field
            is_gi = 1 if p["sub_cat"].lower() == "gi tagged" else 0
            is_organic = 1 if p["sub_cat"].lower() == "organic" else 0
            is_seasonal = 1 if p["category"].lower() == "fruits" else 0
            harvest_season = "summer" if p["category"].lower() == "fruits" else "winter"
            
            # Clean image path
            web_image = clean_image_path(p["image"])
            
            # Use original Sales Price in SQLite schema as price
            cursor.execute("""
                INSERT INTO products (sku, product_title, slug, short_description, category_id, origin_region, unit, price, featured, organic, gi_tagged, seasonal, harvest_season, image, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active')
            """, (p["sku"], p["name"], slug, f"Premium authentic {p['name']} from our farmer cooperatives.", 
                  category_id, "Bihar, India", p["unit"], p["sales_price"], 1, is_organic, is_gi, is_seasonal, harvest_season, web_image))
            
        conn.commit()
        conn.close()
        print("SQLite Database synchronized successfully.")
    except Exception as e:
        print(f"Error during SQLite synchronization: {e}")
        sys.exit(1)
        
    # Generate catalog-data.js for Front-end
    print(f"Generating static JavaScript catalog file: {JS_FILE}...")
    
    cats_js_data = []
    for c in DEFAULT_CATEGORIES:
        cats_js_data.append({
            "id": c["id"],
            "category_name": c["category_name"],
            "slug": c["slug"],
            "parent_id": c["parent_id"],
            "status": bool(c["status"])
        })
        
    prods_js_data = []
    for p in products:
        slug = create_slug(p["name"])
        is_gi = bool(p["sub_cat"].lower() == "gi tagged")
        is_organic = bool(p["sub_cat"].lower() == "organic")
        is_seasonal = bool(p["category"].lower() == "fruits")
        harvest_season = "summer" if p["category"].lower() == "fruits" else "winter"
        
        # Parent category mapping
        cat_slug_mapping = {
            "fruits": "fruits-fresh-produce",
            "grains": "staples-grains",
            "makhana": "superfoods-snacks",
            "sweets": "superfoods-snacks",
            "puja": "puja-essentials-kits",
            "crafts": "handlooms-handicrafts"
        }
        parent_slug = cat_slug_mapping.get(p["category"].lower(), "fruits-fresh-produce")
        web_image = clean_image_path(p["image"])
        stock_num = parse_stock_quantity(p["stock"])
        
        prods_js_data.append({
            "id": slug,
            "title": p["name"],
            "category": parent_slug,
            "isGI": is_gi,
            "isOrganic": is_organic,
            "isSeasonal": is_seasonal,
            "season": harvest_season,
            "price": p["sales_price"],
            "unit": p["unit"],
            "image": web_image,
            "origin": "Bihar, India",
            "popularity": 95,
            "inStock": stock_num > 0,
            "description": f"Premium authentic {p['name']} from our farmer cooperatives.",
            "heritageStory": f"Handpicked and naturally sourced by ZoloFresh farmers. This authentic {p['name']} supports sustainable local livelihood cooperatives."
        })
        
    js_content = f"""// ==========================================
// ZOLOFRESH - Generated Catalog Data (Excel Source)
// Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
// ==========================================

const DEFAULT_CATEGORIES = {json.dumps(cats_js_data, indent=2)};

const DEFAULT_PRODUCTS = {json.dumps(prods_js_data, indent=2)};
"""
    
    try:
        with open(JS_FILE, 'w', encoding='utf-8') as f:
            f.write(js_content)
        print(f"JavaScript catalog successfully exported: {JS_FILE}")
    except Exception as e:
        print(f"Error writing catalog-data.js: {e}")
        sys.exit(1)
        
    print("Catalog sync completed successfully!")


if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] == '--export':
        export_to_excel()
    else:
        import_from_excel()
